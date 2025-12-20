# build_catalog_tistory_best.py
# --------------------------------------------------
# Tistory catalog builder (KO/EN)
# - Language inference supports tokens anywhere in filename (e.g. xxx_en_yyy.yaml, xxx-kr-v2.yml)
# - Excel override + viz rules aligned with GitHub builder
# - KO/EN sorting:
#     * KO: Hangul first (가나다순), then non-Hangul (A–Z)
#     * EN: A–Z
# - Optional client-side search box (name/tags/description)
# - Writes BOTH:
#     tistory/index_ko.html, tistory/index_en.html
#     site/tistory/index_ko.html, site/tistory/index_en.html
# --------------------------------------------------

import os
import re
import html
import json
import yaml

# ===== Optional Excel =====
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# ===== Paths =====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CATALOG_DIR = os.path.join(BASE_DIR, "catalog")

TISTORY_DIR = os.path.join(BASE_DIR, "tistory")
SITE_DIR = os.path.join(BASE_DIR, "site")
SITE_TISTORY_DIR = os.path.join(SITE_DIR, "tistory")

OUT_KO = os.path.join(TISTORY_DIR, "index_ko.html")
OUT_EN = os.path.join(TISTORY_DIR, "index_en.html")
OUT_KO_SITE = os.path.join(SITE_TISTORY_DIR, "index_ko.html")
OUT_EN_SITE = os.path.join(SITE_TISTORY_DIR, "index_en.html")

os.makedirs(TISTORY_DIR, exist_ok=True)
os.makedirs(SITE_TISTORY_DIR, exist_ok=True)

# ===== GitHub Pages base (NO trailing slash) =====
GITHUB_PAGES_BASE = "https://cnucho.github.io/customgpt-catalog"

# ===== Excel index =====
INDEX_XLSX = os.path.join(BASE_DIR, "gpt_index.xlsx")
SHEET_EN_CANDIDATES = ["gpt_index_en", "index_en"]
SHEET_KO_CANDIDATES = ["gpt_index_ko", "index_ko"]

DEFAULT_VIZ1 = "public"     # public | restricted
DEFAULT_VIZ2 = "show_url"   # show_url | hide_url

# ===== Utils =====
def esc(x):
    return html.escape("" if x is None else str(x), quote=True)

def sanitize_id(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9\-_]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "item"

def read_yaml(path: str):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception as e:
        print(f"[SKIP] invalid YAML: {os.path.basename(path)}")
        print(f"       {type(e).__name__}: {e}")
        return None

_LANG_EN_RE = re.compile(r"(^|[_-])(en|eng|english)(?=[_-]|$)")
_LANG_KO_RE = re.compile(r"(^|[_-])(ko|kr|kor|korean)(?=[_-]|$)")

def guess_lang(entry: dict, filename: str) -> str:
    """Infer language from filename token OR YAML fields.

    Token must be a distinct segment separated by '_' or '-', e.g.
      abc_en_title.yaml, abc-kr-v2.yml, en_abc.yaml
    This avoids accidental matches like 'green.yaml' (contains 'en' but not as a token).
    """
    fn = (filename or "").lower()
    base = re.sub(r"\.(yaml|yml)$", "", fn)

    if _LANG_KO_RE.search(base):
        return "ko"
    if _LANG_EN_RE.search(base):
        return "en"

    lang = (entry.get("language") or entry.get("lang") or "").lower().strip()
    if lang in ("ko", "kr", "kor", "korean"):
        return "ko"
    if lang in ("en", "eng", "english"):
        return "en"

    return "en"  # keep legacy default (EN)

# ===== Excel =====
def _load_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = {}
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        rec = dict(zip(header, r))
        raw_key = str(rec.get("gpt_id") or "").strip()
        if not raw_key:
            continue
        key = sanitize_id(raw_key)
        out[key] = rec
        # also store raw (for backward compatibility)
        if raw_key not in out:
            out[raw_key] = rec
    return out

def load_excel_index():
    if not os.path.isfile(INDEX_XLSX) or load_workbook is None:
        return {}, {}

    try:
        wb = load_workbook(INDEX_XLSX, data_only=True)
    except Exception as e:
        print(f"[WARN] Excel load failed: {type(e).__name__}: {e}")
        return {}, {}

    idx_en, idx_ko = {}, {}

    for name in SHEET_EN_CANDIDATES:
        idx_en = _load_sheet(wb, name)
        if idx_en:
            break
    for name in SHEET_KO_CANDIDATES:
        idx_ko = _load_sheet(wb, name)
        if idx_ko:
            break

    if not idx_en:
        print(f"[WARN] Excel EN sheet missing: tried {SHEET_EN_CANDIDATES}")
    if not idx_ko:
        print(f"[WARN] Excel KO sheet missing: tried {SHEET_KO_CANDIDATES}")

    return idx_en, idx_ko

def apply_excel_override(e, idx_en, idx_ko):
    key = e["_id"]
    rec_en = idx_en.get(key)
    rec_ko = idx_ko.get(key)

    # Fallback: match by original gpt_id if present
    if rec_en is None and rec_ko is None:
        raw_gid = str(e.get("gpt_id") or "").strip()
        if raw_gid:
            rec_en = idx_en.get(raw_gid) or idx_en.get(sanitize_id(raw_gid))
            rec_ko = idx_ko.get(raw_gid) or idx_ko.get(sanitize_id(raw_gid))

    if rec_en and rec_en.get("name"):
        e["name_en"] = str(rec_en["name"])
    if rec_ko:
        if rec_ko.get("ko_alias"):
            e["name_ko"] = str(rec_ko["ko_alias"])
        elif rec_ko.get("name"):
            e["name_ko"] = str(rec_ko["name"])

    src = rec_ko or rec_en or {}
    e["viz1"] = str(src.get("viz1", e.get("viz1", DEFAULT_VIZ1)) or DEFAULT_VIZ1)
    e["viz2"] = str(src.get("viz2", e.get("viz2", DEFAULT_VIZ2)) or DEFAULT_VIZ2)

    extra = src.get("extra_fields")
    if extra:
        try:
            obj = json.loads(extra)
            if isinstance(obj, dict):
                e.update(obj)
        except Exception:
            pass

# ===== Display =====
def display_name(e):
    en = (e.get("name_en") or e.get("name") or "").strip()
    ko = (e.get("name_ko") or "").strip()
    if e["_lang"] == "ko":
        return f"{ko} · {en}" if ko and en else (ko or en)
    return f"{en} · {ko}" if en and ko else (en or ko)

def one_line(e, lang):
    return (
        e.get(f"one_line_{lang}")
        or e.get("one_line")
        or e.get("one_line_ko")
        or e.get("one_line_en")
        or ""
    )

def detail_url(e, lang):
    if lang == "ko":
        return f"{GITHUB_PAGES_BASE}/details/{e['_id']}_ko.html"
    return f"{GITHUB_PAGES_BASE}/details/{e['_id']}_en.html"

# ===== Sorting =====
_HANGUL_START_RE = re.compile(r"^\s*[가-힣]")

def _primary_display_ko(e: dict) -> str:
    ko = (e.get("name_ko") or "").strip()
    en = (e.get("name_en") or e.get("name") or "").strip()
    return ko or en or str(e.get("_id", ""))

def sort_key_ko_index(e: dict):
    primary = _primary_display_ko(e)
    if _HANGUL_START_RE.match(primary):
        return (0, primary, str(e.get("_id", "")))
    return (1, primary.lower(), str(e.get("_id", "")))

def sort_key_en_index(e: dict):
    name = (e.get("name_en") or e.get("name") or e.get("_id") or "").strip()
    return (name.lower(), str(e.get("_id", "")))

# ===== Search box =====
def search_box_html(lang: str) -> str:
    placeholder = "이름/태그/설명 검색…" if lang == "ko" else "Search name/tags/description…"
    tip = "Tip: 이름/태그/설명에서 함께 검색됩니다." if lang == "ko" else "Tip: Search matches name, tags, and description."
    return f"""
<div style="margin:12px 0 16px;">
  <input id="q" type="search" placeholder="{placeholder}" oninput="filterList()"
         style="width:100%;max-width:720px;padding:10px 12px;border:1px solid #e5e7eb;border-radius:12px;font-size:14px;"/>
  <div style="margin-top:6px;color:#6b7280;font-size:12px;">{tip}</div>
</div>
<script>
function norm(s){{ return (s||"").toString().toLowerCase(); }}
function filterList(){{
  var q = norm(document.getElementById("q").value);
  var items = document.querySelectorAll(".gpt-item");
  items.forEach(function(it){{
    var hay = norm(it.getAttribute("data-name")) + " " + norm(it.getAttribute("data-tags")) + " " + norm(it.getAttribute("data-one"));
    it.style.display = (q && hay.indexOf(q) === -1) ? "none" : "";
  }});
}}
</script>
""".strip()

# ===== Render =====
def render(title, entries, lang, enable_search=True):
    rows = []
    for i, e in enumerate(entries, start=1):
        restricted = (e.get("viz1") == "restricted")
        show_url = (e.get("viz2") != "hide_url")

        tags = e.get("tags") or []
        if isinstance(tags, str):
            tags_s = tags
        elif isinstance(tags, list):
            tags_s = " ".join(str(x) for x in tags if x is not None)
        else:
            tags_s = str(tags)

        dn = display_name(e)
        ol = one_line(e, lang)

        gpt_btn = ""
        if e.get("url") and show_url:
            label = "GPT 바로가기" if lang == "ko" else "Open GPT"
            gpt_btn = f"<a class='btn' href='{esc(e['url'])}' target='_blank' rel='noopener noreferrer'>{esc(label)}</a>"

        if restricted:
            detail_btn = "<span class='btn restricted'>제한공개</span>" if lang == "ko" else "<span class='btn restricted'>Restricted</span>"
        else:
            label = "상세정보" if lang == "ko" else "Details"
            detail_btn = f"<a class='btn' href='{esc(detail_url(e, lang))}' target='_blank' rel='noopener noreferrer'>{esc(label)}</a>"

        rows.append(f"""
<li class="item gpt-item" data-name="{esc(dn)}" data-tags="{esc(tags_s)}" data-one="{esc(ol)}">
  <div class="title"><strong>{i}. {esc(dn)}</strong></div>
  <div class="meta">{esc(ol)}</div>
  <div class="links">
    {gpt_btn}
    {detail_btn}
  </div>
</li>
""")

    search_html = search_box_html(lang) if enable_search else ""
    switch = ""
    if lang == "ko":
        switch = "<div class='toplinks'><a class='btn small' href='index_en.html'>English</a></div>"
    else:
        switch = "<div class='toplinks'><a class='btn small' href='index_ko.html'>Korean</a></div>"

    return f"""<!doctype html>
<html lang="{ 'ko' if lang=='ko' else 'en' }">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(title)}</title>
<style>
  body {{ margin: 0; padding: 18px; font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif; }}
  .wrap {{ max-width: 980px; margin: 0 auto; }}
  ul {{ list-style:none; padding:0; margin: 0; }}
  .item {{ padding:12px 0; border-bottom:1px solid #ddd; }}
  .meta {{ color:#555; margin:6px 0 10px; line-height:1.5; }}
  .links {{ display:flex; gap:8px; flex-wrap:wrap; }}
  .btn {{ padding:6px 10px; border:1px solid #ccc; border-radius:10px; text-decoration:none; color:#111; background:#fff; font-size: 13px; }}
  .btn:hover {{ background:#f6f6f6; }}
  .btn.small {{ font-size: 12px; padding: 5px 9px; }}
  .restricted {{ background:#fff3cd; color:#7a5a00; border-color:#e5e7eb; cursor: default; }}
  h1 {{ margin: 0 0 8px; font-size: 22px; }}
  .toplinks {{ margin: 8px 0 14px; }}
</style>
</head>
<body>
<div class="wrap">
  <h1>{esc(title)}</h1>
  {switch}
  {search_html}
  <ul>
    {''.join(rows) if rows else ('<li>항목 없음</li>' if lang=='ko' else '<li>No items</li>')}
  </ul>
</div>
</body>
</html>
"""

def build():
    if not os.path.isdir(CATALOG_DIR):
        raise SystemExit(f"[ERROR] catalog dir not found: {CATALOG_DIR}")

    idx_en, idx_ko = load_excel_index()

    all_entries = []
    for fn in sorted(os.listdir(CATALOG_DIR)):
        if not fn.lower().endswith((".yml", ".yaml")):
            continue

        raw = read_yaml(os.path.join(CATALOG_DIR, fn))
        if raw is None or not isinstance(raw, dict):
            continue

        e = raw.get("catalog_entry") or raw.get("gpt") or raw
        if not isinstance(e, dict):
            continue

        e["_lang"] = guess_lang(e, fn)

        raw_id = e.get("_id") or e.get("id") or e.get("gpt_id") or os.path.splitext(fn)[0]
        e["_id"] = sanitize_id(str(raw_id))

        apply_excel_override(e, idx_en, idx_ko)
        all_entries.append(e)

    ko_entries = [e for e in all_entries if e.get("_lang") == "ko"]
    en_entries = [e for e in all_entries if e.get("_lang") == "en"]

    ko_entries.sort(key=sort_key_ko_index)
    en_entries.sort(key=sort_key_en_index)

    ko_html = render("GPT Catalog (KO)", ko_entries, "ko", enable_search=True)
    en_html = render("GPT Catalog (EN)", en_entries, "en", enable_search=True)

    for p, c in [
        (OUT_KO, ko_html), (OUT_EN, en_html),
        (OUT_KO_SITE, ko_html), (OUT_EN_SITE, en_html),
    ]:
        with open(p, "w", encoding="utf-8") as f:
            f.write(c)

    print("[OK] Tistory KO:", len(ko_entries))
    print("[OK] Tistory EN:", len(en_entries))
    print("[OK] wrote:", OUT_KO)
    print("[OK] wrote:", OUT_EN)
    print("[OK] wrote:", OUT_KO_SITE)
    print("[OK] wrote:", OUT_EN_SITE)

if __name__ == "__main__":
    build()
