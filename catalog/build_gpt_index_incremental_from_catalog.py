import os
import yaml
from openpyxl import load_workbook, Workbook

CATALOG_DIR = "catalog"
OUT_XLSX = "gpt_index.xlsx"
SHEET_NAME = "index_base"

DEFAULT_VIZ1 = "public"
DEFAULT_VIZ2 = "show_url"


def read_yaml(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception:
        return None


def extract_entry(raw):
    if not isinstance(raw, dict):
        return None
    if "catalog_entry" in raw and isinstance(raw["catalog_entry"], dict):
        return raw["catalog_entry"]
    if "gpt" in raw and isinstance(raw["gpt"], dict):
        return raw["gpt"]
    return raw


def load_existing_index(path):
    if not os.path.exists(path):
        return None, set()

    wb = load_workbook(path)
    ws = wb[SHEET_NAME]

    existing_ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            existing_ids.add(str(row[0]).strip())

    return wb, existing_ids


def main():
    wb, existing_ids = load_existing_index(OUT_XLSX)

    if wb is None:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["gpt_id", "name", "url", "viz1", "viz2"])
        existing_ids = set()
    else:
        ws = wb[SHEET_NAME]

    added = 0

    for fn in sorted(os.listdir(CATALOG_DIR)):
        if not fn.lower().endswith((".yml", ".yaml")):
            continue

        raw = read_yaml(os.path.join(CATALOG_DIR, fn))
        if raw is None:
            continue

        e = extract_entry(raw)
        if not e:
            continue

        gpt_id = e.get("gpt_id") or e.get("id") or e.get("_id")
        name = e.get("name_en") or e.get("name") or e.get("name_ko")
        url = e.get("url") or ""

        if not gpt_id:
            continue

        gpt_id = str(gpt_id).strip()
        if gpt_id in existing_ids:
            continue  # 핵심: 기존 인덱스는 건드리지 않음

        ws.append([
            gpt_id,
            name or "",
            url,
            DEFAULT_VIZ1,
            DEFAULT_VIZ2,
        ])
        existing_ids.add(gpt_id)
        added += 1

    wb.save(OUT_XLSX)
    print(f"[OK] incremental update complete — added {added} GPT(s)")


if __name__ == "__main__":
    main()
