# build_catalog_github_final.py
# GitHub Pages catalog builder (docs/)
# - BASE_DIR-fixed paths (works regardless of cwd)
# - Excel override with sheet name fallback (gpt_index_en/ko or index_en/ko)
# - Robust YAML parsing (skip invalid YAML, continue)
# - Windows/Dropbox file-lock tolerant clean (retry; if still locked -> rename old folders)
# - Sorting:
#     * KO index: displayed name starts with Hangul first (가나다순), then non-Hangul (A–Z)
#     * EN index: displayed name A–Z
#
# Output:
#   docs/index.html
#   docs/en/index.html
#   docs/details/<id>_ko.html
#   docs/details/<id>_en.html

import os
import re
import html
import json
import shutil
import time
import datetime
from urllib.parse import quote

import yaml

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# ===== Paths (fixed) =====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CATALOG_DIR = os.path.join(BASE_DIR, "catalog")
OUT_DIR = os.path.join(BASE_DIR, "docs")

DETAILS_DIR = os.path.join(OUT_DIR, "details")
EN_DIR = os.path.join(OUT_DIR, "en")

# ===== Excel =====
INDEX_XLSX = os.path.join(BASE_DIR, "gpt_index.xlsx")

# accept both naming conventions
SHEET_EN_CANDIDATES = ["gpt_index_en", "index_en"]
SHEET_KO_CANDIDATES = ["gpt_index_ko", "index_ko"]

# ===== Misc =====
TISTORY_BACK_DEFAULT = "https://skcho.tistory.com/129"
DEFAULT_VIZ1 = "public"     # public | restricted
DEFAULT_VIZ2 = "show_url"   # show_url | hide_url

# ===== Utils =====
def esc(x) -> str:
    return html.escape("" if x is None else str(x), quote=True)

def sanitize_id(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9\-_]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "item"

def read_yaml(path: str):
    try:
        with open(path, "r", encoding="utf-8") as f:
            obj = yaml.safe_load(f)
        return obj or {}
    except Exception as e:
        print(f"[SKIP] invalid YAML: {path}")
        print(f"       {type(e).__name__}: {e}")
        return None

def list_catalog_files():
    if not os.path.isdir(CATALOG_DIR):
        return []
    out = []
    for fn in os.listdir(CATALOG_DIR):
        if fn.lower().endswith((".yml", ".yaml")):
            out.append(fn)
    return sorted(out)

def guess_lang(entry: dict, filename: str) -> str:
    """Infer language from filename and/or YAML fields.

    Supports language tokens anywhere in the filename, as standalone segments
    separated by '_' or '-' (or at start/end), e.g.
      - something_en_title.yaml
      - something-kr-v2.yml
      - en_something.yaml

    This avoids accidental matches like 'green.yaml' -> 'en'.
    """
    fn = (filename or "").lower()
    base = re.sub(r"\.(yaml|yml)$", "", fn)

    # token-based match (safer than substring)
    if re.search(r"(^|[_-])(en|eng|english)(?=[_-]|$)", base):
        return "en"
    if re.search(r"(^|[_-])(ko|kr|kor|korean)(?=[_-]|$)", base):
        return "ko"

    # fallback: explicit fields inside YAML
    lang = (entry.get("language") or entry.get("lang") or "").lower().strip()
    if lang in ("ko", "kr", "kor", "korean"):
        return "ko"
    if lang in ("en", "eng", "english"):
        return "en"

    return "unknown"

# ===== Windows-lock tolerant delete =====
def _rename_away(path: str) -> bool:
    if not os.path.isdir(path):
        return True
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    new_path = f"{path}.old_{ts}"
    try:
        os.rename(path, new_path)
        print(f"[INFO] Could not delete (locked). Renamed: {path} -> {new_path}")
        return True
    except Exception as e:
        print(f"[WARN] Rename failed for locked folder: {path}")
        print(f"       {type(e).__name__}: {e}")
        return False

def safe_rmtree(path: str, retries: int = 8, delay: float = 0.25):
    if not os.path.isdir(path):
        return
    last_err = None
    for _ in range(retries):
        try:
            shutil.rmtree(path)
            return
        except PermissionError as e:
            last_err = e
            time.sleep(delay)
        except OSError as e:
            last_err = e
            time.sleep(delay)
    # still locked -> rename away and continue
    if not _rename_away(path):
        raise last_err

# ===== Excel loading =====
def _load_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = {}
    for r in rows[1:]:
        if not r or r[0] in (None, ""):
            continue
        rec = dict(zip(header, r))
        raw = str(rec.get("gpt_id") or "").strip()
        if not raw:
            continue
        # store by BOTH raw and sanitized id to avoid silent mismatch
        out[raw] = rec
        out[sanitize_id(raw)] = rec
    return out

def load_excel_index():
    if not os.path.isfile(INDEX_XLSX):
        print(f"[INFO] Excel index not found: {INDEX_XLSX} (skip)")
        return {}, {}
    if load_workbook is None:
        print("[WARN] openpyxl not available. Excel overrides disabled.")
        return {}, {}

    try:
        wb = load_workbook(INDEX_XLSX, data_only=True)
    except Exception as e:
        print(f"[WARN] Excel load failed: {type(e).__name__}: {e}")
        return {}, {}

    idx_en, idx_ko = {}, {}
    for name in SHEET_EN_CANDIDATES:
        idx_en = _load_sheet(wb, name)
        if idx_en:
            break
    for name in SHEET_KO_CANDIDATES:
        idx_ko = _load_sheet(wb, name)
        if idx_ko:
            break

    if not idx_en:
        print(f"[WARN] Excel EN sheet missing: tried {SHEET_EN_CANDIDATES} (skip)")
    if not idx_ko:
        print(f"[WARN] Excel KO sheet missing: tried {SHEET_KO_CANDIDATES} (skip)")
    return idx_en, idx_ko

def apply_excel_override(e: dict, index_en: dict, index_ko: dict):
    """
    Match priority:
      1) by sanitized _id
      2) by raw gpt_id (if present)
    """
    key1 = str(e.get("_id") or "").strip()
    key2 = str(e.get("gpt_id") or "").strip()

    rec_en = index_en.get(key1) or (index_en.get(key2) if key2 else None)
    rec_ko = index_ko.get(key1) or (index_ko.get(key2) if key2 else None)

    if rec_en and rec_en.get("name"):
        e["name_en"] = str(rec_en["name"])
    if rec_ko:
        if rec_ko.get("ko_alias"):
            e["name_ko"] = str(rec_ko["ko_alias"])
        elif rec_ko.get("name"):
            e["name_ko"] = str(rec_ko["name"])

    src = rec_ko or rec_en or {}
    if src.get("viz1"):
        e["viz1"] = str(src["viz1"]).strip()
    e["viz1"] = e.get("viz1", DEFAULT_VIZ1) or DEFAULT_VIZ1

    if src.get("viz2"):
        e["viz2"] = str(src["viz2"]).strip()
    e["viz2"] = e.get("viz2", DEFAULT_VIZ2) or DEFAULT_VIZ2

    extra = src.get("extra_fields") if src else None
    if extra:
        try:
            obj = json.loads(extra)
            if isinstance(obj, dict):
                e.update(obj)
        except Exception:
            pass

# ===== Display names =====
def display_name_ko(e: dict) -> str:
    ko = (e.get("name_ko") or "").strip()
    en = (e.get("name_en") or e.get("name") or "").strip()
    if ko and en:
        return f"{esc(ko)} ({esc(en)})"
    return esc(ko or en or e.get("_id",""))

def display_name_en(e: dict) -> str:
    return esc((e.get("name_en") or e.get("name") or e.get("_id") or "").strip())

def _primary_display_ko(e: dict) -> str:
    # for sorting: what user actually sees as the *main* label on KO index
    ko = (e.get("name_ko") or "").strip()
    en = (e.get("name_en") or e.get("name") or "").strip()
    return ko or en or str(e.get("_id",""))

_HANGUL_START_RE = re.compile(r"^\s*[가-힣]")

def sort_key_ko_index(e: dict):
    name = _primary_display_ko(e)
    if _HANGUL_START_RE.match(name):
        # Hangul first, 가나다순 (Unicode order is 가나다순 for Hangul syllables)
        return (0, name, str(e.get("_id","")))
    # then English/others A–Z
    return (1, name.lower(), str(e.get("_id","")))

def sort_key_en_index(e: dict):
    name = (e.get("name_en") or e.get("name") or e.get("_id") or "").strip()
    return (name.lower(), str(e.get("_id","")))

# ===== HTML templates =====
def ko_index_header(total: int) -> str:
    return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>GPT Catalog (KO)</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 24px; }}
    ul {{ list-style: none; padding: 0; }}
    li {{ padding: 12px 0; border-bottom: 1px solid #eee; }}
    .meta {{ color: #555; font-size: 13px; margin-top: 4px; }}
    .topbar {{ display: flex; flex-wrap: wrap; gap: 10px; align-items: center; margin-bottom: 12px; }}
    .pill {{ font-size: 12px; color: #333; background: #f3f3f3; padding: 6px 10px; border-radius: 999px; text-decoration: none; }}
    .btnbar {{ margin-top: 8px; display: flex; flex-wrap: wrap; gap: 8px; align-items:center; }}
    .btn {{ display:inline-block; font-size: 12px; padding: 6px 10px; border-radius: 10px; background: #f6f6f6; color: #222; text-decoration: none; }}
    .btn-small {{ font-size: 11px; padding: 4px 8px; }}
    .btn:hover {{ background: #ededed; }}
    .restricted {{ font-size: 11px; padding: 4px 8px; border-radius: 10px; background: #fff3cd; color: #7a5a00; }}
    .titlelink {{ color: inherit; text-decoration: none; }}
    .titlelink:hover {{ text-decoration: underline; }}
  </style>
</head>
<body>
  <div class="topbar">
    <span class="pill">Total: {total}</span>
    <a class='pill' href='en/index.html'>English</a>
    <a class="pill" href="{esc(TISTORY_BACK_DEFAULT)}" target="_blank" rel="noopener noreferrer">티스토리 목록</a>
  </div>

  <h1>GPT Catalog (KO)</h1>
  <ul id="list">
"""

def en_index_header(total: int) -> str:
    return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>GPT Catalog (EN)</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 24px; }}
    ul {{ list-style: none; padding: 0; }}
    li {{ padding: 12px 0; border-bottom: 1px solid #eee; }}
    .meta {{ color: #555; font-size: 13px; margin-top: 4px; }}
    .topbar {{ display: flex; flex-wrap: wrap; gap: 10px; align-items: center; margin-bottom: 12px; }}
    .pill {{ font-size: 12px; color: #333; background: #f3f3f3; padding: 6px 10px; border-radius: 999px; text-decoration: none; }}
    .btnbar {{ margin-top: 8px; display: flex; flex-wrap: wrap; gap: 8px; align-items:center; }}
    .btn {{ display:inline-block; font-size: 12px; padding: 6px 10px; border-radius: 10px; background: #f6f6f6; color: #222; text-decoration: none; }}
    .btn-small {{ font-size: 11px; padding: 4px 8px; }}
    .btn:hover {{ background: #ededed; }}
    .restricted {{ font-size: 11px; padding: 4px 8px; border-radius: 10px; background: #fff3cd; color: #7a5a00; }}
    .titlelink {{ color: inherit; text-decoration: none; }}
    .titlelink:hover {{ text-decoration: underline; }}
  </style>
</head>
<body>
  <div class="topbar">
    <span class="pill">Total: {total}</span>
    <a class='pill' href='../index.html'>Korean</a>
    <a class="pill" href="{esc(TISTORY_BACK_DEFAULT)}" target="_blank" rel="noopener noreferrer">Tistory</a>
  </div>

  <h1>GPT Catalog (EN)</h1>
  <ul id="list">
"""

def index_footer() -> str:
    return "</ul>\n</body>\n</html>\n"

def make_detail_link(e: dict, lang: str) -> str:
    if lang == "ko":
        return f"details/{e['_id']}_ko.html"
    return f"../details/{e['_id']}_en.html"

def make_back_link_from_detail(lang: str) -> str:
    return "../index.html" if lang == "ko" else "../en/index.html"

def detail_page(e: dict, lang: str) -> str:
    is_ko = (lang == "ko")

    # Title / subtitle (human-facing)
    title_raw = (e.get("name_ko") or "").strip() if is_ko else (e.get("name_en") or e.get("name") or "").strip()
    if not title_raw:
        title_raw = (e.get("name_en") or e.get("name") or e.get("_id") or "").strip()
    subtitle = (e.get("one_line") or "").strip()

    back_href = make_back_link_from_detail(lang)
    back_label = "← 목록" if is_ko else "← Back"

    show_url = (e.get("viz2", DEFAULT_VIZ2) != "hide_url")
    restricted = (e.get("viz1", DEFAULT_VIZ1) == "restricted")
    gpt_url = (e.get("url") or "").strip()

    primary_label = "GPT 바로가기" if is_ko else "Open GPT"
    tistory_label = "티스토리" if is_ko else "Tistory"

    # Lists
    funcs = e.get("functions") or []
    if isinstance(funcs, str):
        funcs = [funcs]
    alias = e.get("alias") or []
    if isinstance(alias, str):
        alias = [alias]
    tags = e.get("tags") or []
    if isinstance(tags, str):
        tags = [tags]

    def chips(items):
        out = []
        for x in items:
            x = ("" if x is None else str(x)).strip()
            if x:
                out.append(f"<span class='chip'>{esc(x)}</span>")
        return "".join(out)

    # Common meta
    created_at = (e.get("created_at") or e.get("created") or "").strip()
    last_updated = (e.get("last_updated") or e.get("updated_at") or "").strip()
    version = (e.get("version") or "").strip()

    # Optional richer fields (show if present)
    category = (e.get("category") or e.get("categories") or "").strip()
    author = (e.get("author") or e.get("creator") or "").strip()
    source = (e.get("source") or e.get("repo") or e.get("github") or "").strip()
    license_ = (e.get("license") or "").strip()
    notes = (e.get("notes") or e.get("note") or "").strip()

    # Potential long text fields -> collapsible
    prompt = e.get("prompt") or e.get("system_prompt") or e.get("instructions") or ""
    description = e.get("description") or e.get("long_description") or ""
    examples = e.get("examples") or e.get("example") or []

    if isinstance(examples, str):
        examples = [examples]

    examples_html = ""
    if examples:
        examples_html = "<ul class='bullets'>" + "".join(f"<li>{esc(x)}</li>" for x in examples if x) + "</ul>"

    func_html = ""
    if funcs:
        func_html = "<ul class='bullets'>" + "".join(f"<li>{esc(x)}</li>" for x in funcs if x) + "</ul>"

    # Button rendering
    primary_btn = ""
    if gpt_url and show_url and not restricted:
        primary_btn = f"<a class='btn primary' href='{esc(gpt_url)}' target='_blank' rel='noopener noreferrer'>{esc(primary_label)}</a>"
    elif restricted:
        primary_btn = f"<span class='btn badge'>{'제한공개' if is_ko else 'Restricted'}</span>"
    elif gpt_url and not show_url:
        primary_btn = f"<span class='btn badge'>{'URL 숨김' if is_ko else 'URL hidden'}</span>"

    notice = ""
    if restricted:
        notice = "<div class='notice'>" + ("제한공개 항목입니다. (URL 비노출)" if is_ko else "This item is restricted (URL hidden).") + "</div>"

    # URL field (human facing)
    url_row = ""
    if gpt_url:
        if show_url and not restricted:
            url_row = f"<div class='row'><div class='k'>URL</div><div class='v'><a href='{esc(gpt_url)}' target='_blank' rel='noopener noreferrer'>{esc(gpt_url)}</a></div></div>"
        else:
            url_row = "<div class='row'><div class='k'>URL</div><div class='v' style='color:#6b7280;'>(hidden)</div></div>"

    # Technical details rows
    tech_rows = [
        ("_id", e.get("_id","")),
        ("gpt_id", e.get("gpt_id","")),
        ("viz1", e.get("viz1", DEFAULT_VIZ1)),
        ("viz2", e.get("viz2", DEFAULT_VIZ2)),
        ("language", e.get("language") or e.get("lang") or lang),
    ]
    tech_html = "".join(
        f"<div class='row'><div class='k'>{esc(k)}</div><div class='v'>{esc(v)}</div></div>"
        for k, v in tech_rows
        if v not in (None, "")
    )

    # Extra fields (show a few common ones if present but not yet shown)
    extra_keys = []
    for k in sorted(e.keys()):
        if k.startswith("_"):
            continue
        if k in {"name_ko","name_en","name","one_line","functions","alias","tags","created_at","created","last_updated","updated_at","version",
                 "category","categories","author","creator","source","repo","github","license","notes","note",
                 "prompt","system_prompt","instructions","description","long_description","examples","example","url","viz1","viz2","language","lang"}:
            continue
        extra_keys.append(k)

    extra_html = ""
    if extra_keys:
        rows = []
        for k in extra_keys[:25]:
            v = e.get(k)
            # keep it short-ish
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)[:500]
            else:
                v = "" if v is None else str(v)
                v = v[:500]
            rows.append(f"<div class='row'><div class='k'>{esc(k)}</div><div class='v'>{esc(v)}</div></div>")
        extra_html = "".join(rows)

    # Build page
    return f"""<!doctype html>
<html lang="{ 'ko' if is_ko else 'en' }">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(title_raw)}</title>
  <style>
    :root {{
      --bg: #f6f7fb;
      --card: #ffffff;
      --text: #111827;
      --muted: #6b7280;
      --border: #e5e7eb;
      --primary: #111827;
      --primaryText: #ffffff;
      --chip: #f3f4f6;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
      background: var(--bg);
      color: var(--text);
    }}
    .wrap {{
      max-width: 980px;
      margin: 0 auto;
      padding: 24px 16px 56px;
    }}
    .card {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 18px;
      padding: 18px;
      box-shadow: 0 6px 22px rgba(17,24,39,0.06);
    }}
    .topbar {{
      display:flex;
      gap: 10px;
      flex-wrap: wrap;
      align-items:center;
      justify-content: space-between;
      margin-bottom: 12px;
    }}
    .leftbar {{
      display:flex;
      gap: 8px;
      flex-wrap: wrap;
      align-items:center;
    }}
    .btn {{
      display:inline-flex;
      align-items:center;
      justify-content:center;
      gap: 8px;
      padding: 10px 14px;
      border-radius: 14px;
      border: 1px solid var(--border);
      text-decoration:none;
      color: var(--text);
      font-size: 13px;
      background: #fff;
    }}
    .btn.primary {{
      background: var(--primary);
      color: var(--primaryText);
      border-color: var(--primary);
      font-weight: 700;
    }}
    .btn.badge {{
      background: #fff3cd;
      color: #7a5a00;
      border-color: var(--border);
      cursor: default;
    }}
    h1 {{
      margin: 0;
      font-size: 30px;
      letter-spacing: -0.02em;
    }}
    .sub {{
      margin: 10px 0 0;
      color: var(--muted);
      line-height: 1.6;
      font-size: 14px;
    }}
    .notice {{
      margin-top: 12px;
      padding: 10px 12px;
      border-radius: 14px;
      border: 1px solid var(--border);
      background: #fff3cd;
      color: #7a5a00;
      font-size: 13px;
    }}
    .section {{
      margin-top: 14px;
    }}
    .h2 {{
      margin: 0 0 8px;
      font-size: 16px;
      letter-spacing: -0.01em;
    }}
    .chips {{
      display:flex;
      gap: 8px;
      flex-wrap: wrap;
      margin-top: 6px;
    }}
    .chip {{
      display:inline-flex;
      align-items:center;
      padding: 6px 10px;
      border-radius: 999px;
      background: var(--chip);
      border: 1px solid var(--border);
      font-size: 12px;
      color: var(--text);
    }}
    .kv {{
      margin-top: 10px;
      display: grid;
      gap: 8px;
    }}
    .row {{
      display:grid;
      grid-template-columns: 160px 1fr;
      gap: 10px;
      padding: 8px 10px;
      border-radius: 12px;
      background: #f9fafb;
      border: 1px solid var(--border);
      overflow-wrap: anywhere;
    }}
    .k {{ color: var(--muted); font-size: 12px; }}
    .v {{ color: var(--text); font-size: 13px; }}
    .bullets {{
      margin: 0;
      padding-left: 18px;
      color: var(--text);
      line-height: 1.65;
      font-size: 14px;
    }}
    details {{
      margin-top: 14px;
      border: 1px solid var(--border);
      border-radius: 16px;
      background: #fff;
      padding: 10px 12px;
    }}
    summary {{
      cursor: pointer;
      font-weight: 700;
      color: var(--text);
      list-style: none;
      outline: none;
    }}
    summary::-webkit-details-marker {{ display:none; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div class="topbar">
        <div class="leftbar">
          <a class="btn" href="{esc(back_href)}">{esc(back_label)}</a>
          {primary_btn}
          <a class="btn" href="{esc(TISTORY_BACK_DEFAULT)}" target="_blank" rel="noopener noreferrer">{esc(tistory_label)}</a>
        </div>
      </div>

      <h1>{esc(title_raw)}</h1>
      {f"<p class='sub'>{esc(subtitle)}</p>" if subtitle else ""}
      {notice}

      <div class="section">
        <div class="h2">{'요약 정보' if is_ko else 'Overview'}</div>
        <div class="kv">
          {url_row}
          {f"<div class='row'><div class='k'>{'버전' if is_ko else 'Version'}</div><div class='v'>{esc(version)}</div></div>" if version else ""}
          {f"<div class='row'><div class='k'>{'생성일' if is_ko else 'Created'}</div><div class='v'>{esc(created_at)}</div></div>" if created_at else ""}
          {f"<div class='row'><div class='k'>{'업데이트' if is_ko else 'Updated'}</div><div class='v'>{esc(last_updated)}</div></div>" if last_updated else ""}
          {f"<div class='row'><div class='k'>{'카테고리' if is_ko else 'Category'}</div><div class='v'>{esc(category)}</div></div>" if category else ""}
          {f"<div class='row'><div class='k'>{'작성자' if is_ko else 'Author'}</div><div class='v'>{esc(author)}</div></div>" if author else ""}
          {f"<div class='row'><div class='k'>{'라이선스' if is_ko else 'License'}</div><div class='v'>{esc(license_)}</div></div>" if license_ else ""}
          {f"<div class='row'><div class='k'>{'소스' if is_ko else 'Source'}</div><div class='v'>{esc(source)}</div></div>" if source else ""}
        </div>

        {f"<div class='chips'>{chips(tags)}</div>" if tags else ""}
        {f"<div class='chips'>{chips(alias)}</div>" if alias else ""}
        {f"<p class='sub'>{esc(notes)}</p>" if notes else ""}
      </div>

      <div class="section">
        <div class="h2">{'핵심 기능' if is_ko else 'Key functions'}</div>
        {func_html if func_html else ("<p class='sub'>" + ("등록된 기능이 없습니다." if is_ko else "No functions listed.") + "</p>")}
      </div>

      {f"<div class='section'><div class='h2'>{'설명' if is_ko else 'Description'}</div><p class='sub'>{esc(description)}</p></div>" if description else ""}

      {f"<div class='section'><div class='h2'>{'예시' if is_ko else 'Examples'}</div>{examples_html}</div>" if examples_html else ""}

      <details>
        <summary>{'기술 정보' if is_ko else 'Technical details'}</summary>
        <div class="kv">{tech_html}</div>
      </details>

      {f"<details><summary>{'기타 필드' if is_ko else 'Other fields'}</summary><div class='kv'>{extra_html}</div></details>" if extra_html else ""}

      {f"<details><summary>{'프롬프트/지침' if is_ko else 'Prompt / Instructions'}</summary><div class='kv'><div class='row'><div class='k'>text</div><div class='v'>{esc(prompt)}</div></div></div></details>" if prompt else ""}

    </div>
  </div>
</body>
</html>
"""


def build(clean: bool = True):
    # clean output dirs
    if clean:
        safe_rmtree(DETAILS_DIR)
        safe_rmtree(EN_DIR)
        if os.path.isfile(os.path.join(OUT_DIR, "index.html")):
            try:
                os.remove(os.path.join(OUT_DIR, "index.html"))
            except Exception:
                pass

    os.makedirs(DETAILS_DIR, exist_ok=True)
    os.makedirs(EN_DIR, exist_ok=True)

    index_en, index_ko = load_excel_index()

    all_entries = []
    unknown_files = []

    for fn in list_catalog_files():
        path = os.path.join(CATALOG_DIR, fn)
        raw = read_yaml(path)
        if raw is None or not isinstance(raw, dict):
            continue

        # allow wrapper keys
        e = raw.get("catalog_entry") or raw.get("gpt") or raw

        raw_id = e.get("_id") or e.get("id") or e.get("gpt_id") or os.path.splitext(fn)[0]
        e["_id"] = sanitize_id(str(raw_id))

        e["_lang"] = guess_lang(e, fn)
        if e["_lang"] == "unknown":
            unknown_files.append((fn, e.get("_id","")))

        apply_excel_override(e, index_en, index_ko)
        all_entries.append(e)

    ko_entries = [e for e in all_entries if e["_lang"] == "ko"]
    en_entries = [e for e in all_entries if e["_lang"] == "en"]
    unknown_entries = [e for e in all_entries if e["_lang"] == "unknown"]

    # Sort (requested)
    ko_entries.sort(key=sort_key_ko_index)
    en_entries.sort(key=sort_key_en_index)

    print(f"Loaded total: {len(all_entries)} | KO: {len(ko_entries)} | EN: {len(en_entries)}")
    if unknown_entries:
        print(f"[WARN] Unknown-language catalogs skipped: {len(unknown_entries)}")
        for fn, _id in unknown_files[:50]:
            print(f"       - {fn} (id={_id})")
        if len(unknown_files) > 50:
            print(f"       ... (+{len(unknown_files)-50} more)")

    # write KO index
    root_index = os.path.join(OUT_DIR, "index.html")
    with open(root_index, "w", encoding="utf-8") as f:
        f.write(ko_index_header(len(ko_entries)))
        for i, e in enumerate(ko_entries, start=1):
            name = display_name_ko(e)
            one = esc(e.get("one_line", ""))
            tags_raw = e.get("tags", [])
            if isinstance(tags_raw, list):
                tags = esc(" ".join(map(str, tags_raw)))
            else:
                tags = esc(str(tags_raw))
            gpt_url = esc(e.get("url", ""))
            detail_href = make_detail_link(e, "ko")

            restricted = (e.get("viz1", DEFAULT_VIZ1) == "restricted")
            show_url = (e.get("viz2", DEFAULT_VIZ2) != "hide_url")

            if restricted:
                detail_btn = "<span class='restricted'>제한공개</span>"
                title_href = "#"
            else:
                detail_btn = f"<a class='btn btn-small' href='{detail_href}'>상세보기</a>"
                title_href = detail_href

            url_btn = ""
            if gpt_url and show_url:
                url_btn = (f"<a class='btn btn-small' href='{gpt_url}' target='_blank' "
                           f"rel='noopener noreferrer'>GPT 바로가기</a>")

            f.write(
                f"<li data-name='{esc(_primary_display_ko(e))}' data-tags='{tags}' data-one='{one}'>"
                f"<div class='row'><a class='titlelink' href='{title_href}'><strong>{i}. {name}</strong></a></div>"
                f"<div class='meta'>{one}</div>"
                f"<div class='btnbar'>{url_btn}{detail_btn}</div></li>"
            )
        f.write(index_footer())

    # write EN index
    en_index = os.path.join(EN_DIR, "index.html")
    with open(en_index, "w", encoding="utf-8") as f:
        f.write(en_index_header(len(en_entries)))
        for i, e in enumerate(en_entries, start=1):
            name = display_name_en(e)
            one = esc(e.get("one_line", ""))
            tags_raw = e.get("tags", [])
            if isinstance(tags_raw, list):
                tags = esc(" ".join(map(str, tags_raw)))
            else:
                tags = esc(str(tags_raw))
            gpt_url = esc(e.get("url", ""))
            detail_href = make_detail_link(e, "en")

            restricted = (e.get("viz1", DEFAULT_VIZ1) == "restricted")
            show_url = (e.get("viz2", DEFAULT_VIZ2) != "hide_url")

            if restricted:
                detail_btn = "<span class='restricted'>Restricted</span>"
                title_href = "#"
            else:
                detail_btn = f"<a class='btn btn-small' href='{detail_href}'>View details</a>"
                title_href = detail_href

            url_btn = ""
            if gpt_url and show_url:
                url_btn = (f"<a class='btn btn-small' href='{gpt_url}' target='_blank' "
                           f"rel='noopener noreferrer'>Open GPT</a>")

            f.write(
                f"<li data-name='{esc(e.get('name_en') or e.get('name') or e.get('_id') or '')}' "
                f"data-tags='{tags}' data-one='{one}'>"
                f"<div class='row'><a class='titlelink' href='{title_href}'><strong>{i}. {name}</strong></a></div>"
                f"<div class='meta'>{one}</div>"
                f"<div class='btnbar'>{url_btn}{detail_btn}</div></li>"
            )
        f.write(index_footer())

    # details
    for e in ko_entries:
        out = os.path.join(DETAILS_DIR, f"{e['_id']}_ko.html")
        with open(out, "w", encoding="utf-8") as fp:
            fp.write(detail_page(e, "ko"))

    for e in en_entries:
        out = os.path.join(DETAILS_DIR, f"{e['_id']}_en.html")
        with open(out, "w", encoding="utf-8") as fp:
            fp.write(detail_page(e, "en"))

    print(f"[OK] GitHub docs written: {OUT_DIR}")
    print(f"[OK] Details: {DETAILS_DIR}")

if __name__ == "__main__":
    build(clean=True)
