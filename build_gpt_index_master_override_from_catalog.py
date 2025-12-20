import os
import yaml
from datetime import date
from openpyxl import load_workbook, Workbook

CATALOG_DIR = "catalog"
OUT_XLSX = "gpt_index.xlsx"

SHEET_KO = "gpt_index_ko"
SHEET_EN = "gpt_index_en"

# 운영 기본값(새 엔트리 추가 시에만 적용)
DEFAULT_VISIBILITY = "public"   # e.g., public / internal / private
DEFAULT_SHOW_URL = "show_url"   # e.g., show_url / hide_url

# -----------------------
# YAML helpers
# -----------------------
def read_yaml(path: str):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception:
        return None

def extract_entry(raw):
    if not isinstance(raw, dict):
        return None
    if isinstance(raw.get("catalog_entry"), dict):
        return raw["catalog_entry"]
    if isinstance(raw.get("gpt"), dict):
        return raw["gpt"]
    return raw

def norm_str(x):
    if x is None:
        return ""
    return str(x).strip()

# -----------------------
# Excel helpers
# -----------------------
def ensure_sheet_with_headers(wb, sheet_name, required_headers):
    """
    - 시트 없으면 생성 + 헤더 작성
    - 시트 있으면 헤더를 읽고, 누락된 required_headers는 오른쪽에 추가
    - 헤더가 아예 없으면 1행에 required_headers 작성
    """
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(required_headers)
        return ws

    # 헤더 읽기
    if ws.max_row < 1:
        ws.append(required_headers)
        return ws

    header_row = [c.value for c in ws[1]]
    if all(v is None for v in header_row):
        ws.delete_rows(1, 1)
        ws.append(required_headers)
        return ws

    # 누락 컬럼 추가
    existing = [norm_str(v) for v in header_row]
    for h in required_headers:
        if h not in existing:
            ws.cell(row=1, column=ws.max_column + 1, value=h)
            existing.append(h)

    return ws

def header_map(ws):
    """header -> col_index (1-based)"""
    hm = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        key = norm_str(cell.value)
        if key:
            hm[key] = idx
    return hm

def find_row_by_id(ws, hm, gpt_id):
    """Return row index (>=2) if found, else None"""
    id_col = hm.get("gpt_id")
    if not id_col:
        return None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v is None:
            continue
        if norm_str(v) == gpt_id:
            return r
    return None

def get_cell(ws, r, hm, key):
    c = hm.get(key)
    if not c:
        return ""
    return norm_str(ws.cell(row=r, column=c).value)

def set_cell(ws, r, hm, key, value):
    c = hm.get(key)
    if not c:
        return
    ws.cell(row=r, column=c, value=value)

# -----------------------
# Schema (manual vs catalog snapshot vs status)
# -----------------------
KO_HEADERS = [
    # key
    "gpt_id",
    # manual (override)
    "display_name_ko",
    "alias_ko",
    "keywords_ko",
    "visibility",
    "show_url",
    "url_override",
    # catalog snapshot (auto)
    "name_ko__catalog",
    "name_en__catalog",
    "url__catalog",
    "version__catalog",
    "last_updated__catalog",
    # status (auto)
    "present_in_catalog",
    "last_seen",
    "changed",
    "changed_fields",
]

EN_HEADERS = [
    "gpt_id",
    # manual (override)
    "display_name_en",
    "keywords_en",
    "visibility",
    "show_url",
    "url_override",
    # catalog snapshot (auto)
    "name_en__catalog",
    "name_ko__catalog",
    "url__catalog",
    "version__catalog",
    "last_updated__catalog",
    # status (auto)
    "present_in_catalog",
    "last_seen",
    "changed",
    "changed_fields",
]

CATALOG_FIELDS = [
    "name_ko__catalog",
    "name_en__catalog",
    "url__catalog",
    "version__catalog",
    "last_updated__catalog",
]

def catalog_snapshot_from_entry(e: dict):
    # 여러 카탈로그 파일 변형을 최대한 흡수
    name_ko = norm_str(e.get("name_ko") or e.get("name") or "")
    name_en = norm_str(e.get("name_en") or e.get("name") or "")
    url = norm_str(e.get("url") or "")
    version = norm_str(e.get("version") or e.get("ver") or "")
    last_updated = norm_str(e.get("last_updated") or e.get("updated") or "")

    return {
        "name_ko__catalog": name_ko,
        "name_en__catalog": name_en,
        "url__catalog": url,
        "version__catalog": version,
        "last_updated__catalog": last_updated,
    }

def update_row(ws, hm, row_idx, snap: dict, today_iso: str, is_new: bool):
    """
    정책:
    - manual(override) 컬럼은 절대 덮어쓰지 않음
      단, 신규 엔트리(is_new=True)일 때만 기본값/시드값을 채움
    - catalog snapshot 컬럼은 매번 갱신
    - 변경 감지: 이전 snapshot vs 새 snapshot 비교
      * diff 있으면 changed=Y, changed_fields에 필드명 콤마로 기록
      * diff 없으면 changed는 유지(기존 Y를 N으로 강제로 내리지 않음)
    - present_in_catalog=Y, last_seen=today
    """
    # 1) 신규 엔트리면 manual seed
    if is_new:
        # display name seed (빈값으로 남기고 싶으면 아래 2줄을 주석처리)
        if "display_name_ko" in hm:
            set_cell(ws, row_idx, hm, "display_name_ko", snap.get("name_ko__catalog", ""))
        if "display_name_en" in hm:
            set_cell(ws, row_idx, hm, "display_name_en", snap.get("name_en__catalog", ""))

        set_cell(ws, row_idx, hm, "visibility", DEFAULT_VISIBILITY)
        set_cell(ws, row_idx, hm, "show_url", DEFAULT_SHOW_URL)
        # url_override는 기본 공란

    # 2) 변경 감지 (이전 snapshot 읽고 비교)
    changed_fields = []
    for f in CATALOG_FIELDS:
        oldv = get_cell(ws, row_idx, hm, f)
        newv = norm_str(snap.get(f, ""))
        if oldv != newv:
            changed_fields.append(f)

    # 3) snapshot 갱신
    for k, v in snap.items():
        set_cell(ws, row_idx, hm, k, v)

    # 4) status 갱신
    set_cell(ws, row_idx, hm, "present_in_catalog", "Y")
    set_cell(ws, row_idx, hm, "last_seen", today_iso)

    if changed_fields:
        set_cell(ws, row_idx, hm, "changed", "Y")
        set_cell(ws, row_idx, hm, "changed_fields", ",".join(changed_fields))

def mark_missing(ws, hm, seen_ids: set):
    """
    catalog에 없는 gpt_id는 삭제하지 않고 present_in_catalog=N으로 표시.
    """
    id_col = hm.get("gpt_id")
    if not id_col:
        return 0
    missing = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v is None:
            continue
        gid = norm_str(v)
        if not gid:
            continue
        if gid not in seen_ids:
            set_cell(ws, r, hm, "present_in_catalog", "N")
            missing += 1
    return missing

def main():
    today_iso = date.today().isoformat()

    # workbook load / create
    if os.path.exists(OUT_XLSX):
        wb = load_workbook(OUT_XLSX)
        # 기본 Sheet가 있을 수 있으니 놔둬도 되지만, 여기서는 그대로 둠
    else:
        wb = Workbook()
        # 기본 시트 제거(원치 않으면)
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])

    ws_ko = ensure_sheet_with_headers(wb, SHEET_KO, KO_HEADERS)
    ws_en = ensure_sheet_with_headers(wb, SHEET_EN, EN_HEADERS)

    hm_ko = header_map(ws_ko)
    hm_en = header_map(ws_en)

    seen = set()
    added_ko = added_en = 0
    updated_ko = updated_en = 0

    for fn in sorted(os.listdir(CATALOG_DIR)):
        if not fn.lower().endswith((".yml", ".yaml")):
            continue

        raw = read_yaml(os.path.join(CATALOG_DIR, fn))
        if raw is None:
            continue

        e = extract_entry(raw)
        if not isinstance(e, dict) or not e:
            continue

        gpt_id = norm_str(e.get("gpt_id") or e.get("id") or e.get("_id"))
        if not gpt_id:
            continue

        snap = catalog_snapshot_from_entry(e)
        seen.add(gpt_id)

        # ---- KO sheet upsert
        r_ko = find_row_by_id(ws_ko, hm_ko, gpt_id)
        if r_ko is None:
            # append new row with gpt_id only; rest filled by update_row
            ws_ko.append([gpt_id] + [""] * (len(KO_HEADERS) - 1))
            hm_ko = header_map(ws_ko)  # in case cols changed
            r_ko = ws_ko.max_row
            update_row(ws_ko, hm_ko, r_ko, snap, today_iso, is_new=True)
            added_ko += 1
        else:
            update_row(ws_ko, hm_ko, r_ko, snap, today_iso, is_new=False)
            updated_ko += 1

        # ---- EN sheet upsert
        r_en = find_row_by_id(ws_en, hm_en, gpt_id)
        if r_en is None:
            ws_en.append([gpt_id] + [""] * (len(EN_HEADERS) - 1))
            hm_en = header_map(ws_en)
            r_en = ws_en.max_row
            update_row(ws_en, hm_en, r_en, snap, today_iso, is_new=True)
            added_en += 1
        else:
            update_row(ws_en, hm_en, r_en, snap, today_iso, is_new=False)
            updated_en += 1

    # mark missing (do not delete)
    missing_ko = mark_missing(ws_ko, hm_ko, seen)
    missing_en = mark_missing(ws_en, hm_en, seen)

    wb.save(OUT_XLSX)
    print(
        f"[OK] master-override index updated: "
        f"KO added={added_ko}, updated={updated_ko}, missing_marked={missing_ko}; "
        f"EN added={added_en}, updated={updated_en}, missing_marked={missing_en} "
        f"-> {OUT_XLSX}"
    )


if __name__ == "__main__":
    main()
