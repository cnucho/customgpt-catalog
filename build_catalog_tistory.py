# build_tistory_catalog.nosearch.py
# Tistory catalog builder
# - ID / Excel override / viz rules aligned with GitHub builder

import os
import re
import html
import json
import yaml
from urllib.parse import quote

# ===== Optional Excel =====
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# ===== Paths =====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CATALOG_DIR = os.path.join(BASE_DIR, "catalog")
TISTORY_DIR = os.path.join(BASE_DIR, "tistory")
SITE_DIR = os.path.join(BASE_DIR, "site")
SITE_TISTORY_DIR = os.path.join(SITE_DIR, "tistory")

OUT_KO = os.path.join(TISTORY_DIR, "index_ko.html")
OUT_EN = os.path.join(TISTORY_DIR, "index_en.html")
OUT_KO_SITE = os.path.join(SITE_TISTORY_DIR, "index_ko.html")
OUT_EN_SITE = os.path.join(SITE_TISTORY_DIR, "index_en.html")

os.makedirs(TISTORY_DIR, exist_ok=True)
os.makedirs(SITE_TISTORY_DIR, exist_ok=True)

# ===== GitHub Pages base (NO trailing slash) =====
GITHUB_PAGES_BASE = "https://cnucho.github.io/customgpt-catalog"

# ===== Excel index =====
INDEX_XLSX = os.path.join(BASE_DIR, "gpt_index.xlsx")
SHEET_EN = "index_en"
SHEET_KO = "index_ko"

DEFAULT_VIZ1 = "public"     # public | restricted
DEFAULT_VIZ2 = "show_url"   # show_url | hide_url

# ===== Utils =====
def esc(x):
    return html.escape("" if x is None else str(x), quote=True)

def sanitize_id(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9\-_]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "item"

def read_yaml(path: str):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception as e:
        print(f"[SKIP] invalid YAML: {os.path.basename(path)}")
        print(f"       {e}")
        return None

def guess_lang(entry: dict, filename: str) -> str:
    fn = filename.lower()
    if any(x in fn for x in ("_ko", "-ko", "_kr", "-kr", "_kor", "-kor")):
        return "ko"
    if any(x in fn for x in ("_en", "-en")):
        return "en"

    lang = (entry.get("language") or entry.get("lang") or "").lower()
    if lang in ("ko", "kr", "kor", "korean"):
        return "ko"
    return "en"   # default EN (same as current behavior)

# ===== Excel =====
def load_excel_index():
    if not os.path.isfile(INDEX_XLSX) or load_workbook is None:
        return {}, {}

    def _load(sheet):
        try:
            wb = load_workbook(INDEX_XLSX, data_only=True)
            if sheet not in wb.sheetnames:
                return {}
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            header = [str(h).strip() for h in rows[0]]
            out = {}
            for r in rows[1:]:
                if not r or not r[0]:
                    continue
                rec = dict(zip(header, r))
                raw_key = str(rec.get("gpt_id") or "").strip()
                if not raw_key:
                    continue
                key = sanitize_id(raw_key)
                out[key] = rec
                if raw_key not in out:
                    out[raw_key] = rec
            return out
        except Exception:
            return {}

    return _load(SHEET_EN), _load(SHEET_KO)

def apply_excel_override(e, idx_en, idx_ko):
    key = e["_id"]
    rec_en = idx_en.get(key)
    rec_ko = idx_ko.get(key)

    # Fallback: match by original gpt_id if present
    if (rec_en is None and rec_ko is None):
        raw_gid = str(e.get("gpt_id") or "").strip()
        if raw_gid:
            rec_en = idx_en.get(raw_gid) or idx_en.get(sanitize_id(raw_gid))
            rec_ko = idx_ko.get(raw_gid) or idx_ko.get(sanitize_id(raw_gid))

    if rec_en and rec_en.get("name"):
        e["name_en"] = rec_en["name"]
    if rec_ko:
        if rec_ko.get("ko_alias"):
            e["name_ko"] = rec_ko["ko_alias"]
        elif rec_ko.get("name"):
            e["name_ko"] = rec_ko["name"]

    src = rec_ko or rec_en or {}
    e["viz1"] = str(src.get("viz1", e.get("viz1", DEFAULT_VIZ1)))
    e["viz2"] = str(src.get("viz2", e.get("viz2", DEFAULT_VIZ2)))

    extra = src.get("extra_fields")
    if extra:
        try:
            e.update(json.loads(extra))
        except Exception:
            pass

# ===== Display =====
def display_name(e):
    en = e.get("name_en") or e.get("name") or ""
    ko = e.get("name_ko") or ""
    if e["_lang"] == "ko":
        return f"{ko} · {en}" if ko and en else (ko or en)
    return f"{en} · {ko}" if en and ko else (en or ko)

def one_line(e, lang):
    return (
        e.get(f"one_line_{lang}")
        or e.get("one_line")
        or e.get("one_line_ko")
        or e.get("one_line_en")
        or ""
    )

def detail_url(e, lang):
    if lang == "ko":
        return f"{GITHUB_PAGES_BASE}/details/{e['_id']}_ko.html"
    return f"{GITHUB_PAGES_BASE}/details/{e['_id']}_en.html"

# ===== Render =====
def render(title, entries, lang):
    rows = []
    for i, e in enumerate(entries, start=1):
        restricted = (e.get("viz1") == "restricted")
        show_url = (e.get("viz2") != "hide_url")

        gpt_btn = ""
        if e.get("url") and show_url:
            gpt_btn = f"<a class='btn' href='{esc(e['url'])}' target='_blank'>GPT 바로가기</a>"

        if restricted:
            detail_btn = "<span class='btn restricted'>제한공개</span>"
        else:
            detail_btn = f"<a class='btn' href='{esc(detail_url(e, lang))}' target='_blank'>상세정보</a>"

        rows.append(f"""
<li class="item">
  <div class="title"><strong>{i}. {esc(display_name(e))}</strong></div>
  <div class="meta">{esc(one_line(e, lang))}</div>
  <div class="links">
    {gpt_btn}
    {detail_btn}
  </div>
</li>
""")

    return f"""
<div class="wrap">
<style>
  .wrap {{ font-family:system-ui; }}
  ul {{ list-style:none; padding:0; }}
  .item {{ padding:12px 0; border-bottom:1px solid #ddd; }}
  .meta {{ color:#555; margin:6px 0 10px; }}
  .links {{ display:flex; gap:8px; flex-wrap:wrap; }}
  .btn {{ padding:6px 10px; border:1px solid #ccc; border-radius:8px; text-decoration:none; }}
  .restricted {{ background:#fff3cd; color:#7a5a00; }}
</style>

<h3>{esc(title)}</h3>
<ul>
{''.join(rows) if rows else '<li>항목 없음</li>'}
</ul>
</div>
"""

# =========================
# Build
# =========================
idx_en, idx_ko = load_excel_index()

all_entries = []
for fn in sorted(os.listdir(CATALOG_DIR)):
    if not fn.lower().endswith((".yml", ".yaml")):
        continue

    raw = read_yaml(os.path.join(CATALOG_DIR, fn))
    if raw is None or not isinstance(raw, dict):
        continue

    e = raw.get("catalog_entry") or raw.get("gpt") or raw
    e["_lang"] = guess_lang(e, fn)

    raw_id = e.get("_id") or e.get("id") or e.get("gpt_id") or os.path.splitext(fn)[0]
    e["_id"] = sanitize_id(raw_id)

    apply_excel_override(e, idx_en, idx_ko)
    all_entries.append(e)

ko_entries = [e for e in all_entries if e["_lang"] == "ko"]
en_entries = [e for e in all_entries if e["_lang"] == "en"]

ko_html = render("GPT Catalog (KO)", ko_entries, "ko")
en_html = render("GPT Catalog (EN)", en_entries, "en")

for p, c in [
    (OUT_KO, ko_html), (OUT_EN, en_html),
    (OUT_KO_SITE, ko_html), (OUT_EN_SITE, en_html),
]:
    with open(p, "w", encoding="utf-8") as f:
        f.write(c)

print("[OK] Tistory KO:", len(ko_entries))
print("[OK] Tistory EN:", len(en_entries))
