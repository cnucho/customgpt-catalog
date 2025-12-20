# build_catalog_github_pretty_fixed.py
# GitHub Pages catalog builder (docs/) with:
# - Clean card UI for index + detail
# - Built-in client-side search (name/tags/one_line)
# - KO sort: Hangul-first 가나다순, then English A–Z
# - EN sort: A–Z
# - BASE_DIR-fixed paths
# - Excel override (sheet name fallback)
# - Windows/Dropbox file-lock tolerant clean (retry; then rename away)

import os
import re
import html
import json
import shutil
import time
import datetime

import yaml

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# ===== Paths (fixed) =====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CATALOG_DIR = os.path.join(BASE_DIR, "catalog")
OUT_DIR = os.path.join(BASE_DIR, "docs")
DETAILS_DIR = os.path.join(OUT_DIR, "details")
EN_DIR = os.path.join(OUT_DIR, "en")

# ===== Excel =====
INDEX_XLSX = os.path.join(BASE_DIR, "gpt_index.xlsx")
SHEET_EN_CANDIDATES = ["gpt_index_en", "index_en"]
SHEET_KO_CANDIDATES = ["gpt_index_ko", "index_ko"]

# ===== Config =====
TISTORY_BACK_DEFAULT = "https://skcho.tistory.com/129"
DEFAULT_VIZ1 = "public"     # public | restricted
DEFAULT_VIZ2 = "show_url"   # show_url | hide_url

# ===== Utils =====
def esc(x) -> str:
    return html.escape("" if x is None else str(x), quote=True)

def sanitize_id(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9\-_]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "item"

def read_yaml(path: str):
    try:
        with open(path, "r", encoding="utf-8") as f:
            obj = yaml.safe_load(f)
        return obj or {}
    except Exception as e:
        print(f"[SKIP] invalid YAML: {path}")
        print(f"       {type(e).__name__}: {e}")
        return None

def list_catalog_files():
    if not os.path.isdir(CATALOG_DIR):
        return []
    return sorted([fn for fn in os.listdir(CATALOG_DIR) if fn.lower().endswith((".yml", ".yaml"))])

def guess_lang(entry: dict, filename: str) -> str:
    fn = filename.lower()
    base = re.sub(r"\.(yaml|yml)$", "", fn)

    if base.endswith("_en") or base.endswith("-en") or "_en_" in base or "-en-" in base:
        return "en"
    if base.endswith(("_ko", "_kr", "_kor")) or base.endswith(("-ko", "-kr", "-kor")) or any(tok in base for tok in ("_ko_", "_kr_", "_kor_", "-ko-", "-kr-", "-kor-")):
        return "ko"

    lang = (entry.get("language") or entry.get("lang") or "").lower().strip()
    if lang in ("ko", "kr", "kor", "korean"):
        return "ko"
    if lang in ("en", "eng", "english"):
        return "en"
    return "unknown"

# ===== Windows-lock tolerant delete =====
def _rename_away(path: str) -> bool:
    if not os.path.isdir(path):
        return True
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    new_path = f"{path}.old_{ts}"
    try:
        os.rename(path, new_path)
        print(f"[INFO] Could not delete (locked). Renamed: {path} -> {new_path}")
        return True
    except Exception as e:
        print(f"[WARN] Rename failed for locked folder: {path}")
        print(f"       {type(e).__name__}: {e}")
        return False

def safe_rmtree(path: str, retries: int = 8, delay: float = 0.25):
    if not os.path.isdir(path):
        return
    last_err = None
    for _ in range(retries):
        try:
            shutil.rmtree(path)
            return
        except (PermissionError, OSError) as e:
            last_err = e
            time.sleep(delay)
    if not _rename_away(path):
        raise last_err

# ===== Excel loading =====
def _load_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = {}
    for r in rows[1:]:
        if not r or r[0] in (None, ""):
            continue
        rec = dict(zip(header, r))
        raw = str(rec.get("gpt_id") or "").strip()
        if not raw:
            continue
        out[raw] = rec
        out[sanitize_id(raw)] = rec
    return out

def load_excel_index():
    if not os.path.isfile(INDEX_XLSX):
        print(f"[INFO] Excel index not found: {INDEX_XLSX} (skip)")
        return {}, {}
    if load_workbook is None:
        print("[WARN] openpyxl not available. Excel overrides disabled.")
        return {}, {}
    try:
        wb = load_workbook(INDEX_XLSX, data_only=True)
    except Exception as e:
        print(f"[WARN] Excel load failed: {type(e).__name__}: {e}")
        return {}, {}

    idx_en, idx_ko = {}, {}
    for name in SHEET_EN_CANDIDATES:
        idx_en = _load_sheet(wb, name)
        if idx_en:
            break
    for name in SHEET_KO_CANDIDATES:
        idx_ko = _load_sheet(wb, name)
        if idx_ko:
            break

    if not idx_en:
        print(f"[WARN] Excel EN sheet missing: tried {SHEET_EN_CANDIDATES} (skip)")
    if not idx_ko:
        print(f"[WARN] Excel KO sheet missing: tried {SHEET_KO_CANDIDATES} (skip)")
    return idx_en, idx_ko

def apply_excel_override(e: dict, index_en: dict, index_ko: dict):
    key1 = str(e.get("_id") or "").strip()
    key2 = str(e.get("gpt_id") or "").strip()
    rec_en = index_en.get(key1) or (index_en.get(key2) if key2 else None)
    rec_ko = index_ko.get(key1) or (index_ko.get(key2) if key2 else None)

    if rec_en and rec_en.get("name"):
        e["name_en"] = str(rec_en["name"])
    if rec_ko:
        if rec_ko.get("ko_alias"):
            e["name_ko"] = str(rec_ko["ko_alias"])
        elif rec_ko.get("name"):
            e["name_ko"] = str(rec_ko["name"])

    src = rec_ko or rec_en or {}
    if src.get("viz1"):
        e["viz1"] = str(src["viz1"]).strip()
    e["viz1"] = e.get("viz1", DEFAULT_VIZ1) or DEFAULT_VIZ1

    if src.get("viz2"):
        e["viz2"] = str(src["viz2"]).strip()
    e["viz2"] = e.get("viz2", DEFAULT_VIZ2) or DEFAULT_VIZ2

    extra = src.get("extra_fields") if src else None
    if extra:
        try:
            obj = json.loads(extra)
            if isinstance(obj, dict):
                e.update(obj)
        except Exception:
            pass

# ===== Display & sorting =====
def display_name_ko(e: dict) -> str:
    ko = (e.get("name_ko") or "").strip()
    en = (e.get("name_en") or e.get("name") or "").strip()
    if ko and en:
        return f"{esc(ko)} ({esc(en)})"
    return esc(ko or en or e.get("_id",""))

def display_name_en(e: dict) -> str:
    return esc((e.get("name_en") or e.get("name") or e.get("_id") or "").strip())

def _primary_display_ko(e: dict) -> str:
    ko = (e.get("name_ko") or "").strip()
    en = (e.get("name_en") or e.get("name") or "").strip()
    return ko or en or str(e.get("_id",""))

_HANGUL_START_RE = re.compile(r"^\s*[가-힣]")

def sort_key_ko_index(e: dict):
    name = _primary_display_ko(e)
    if _HANGUL_START_RE.match(name):
        return (0, name, str(e.get("_id","")))
    return (1, name.lower(), str(e.get("_id","")))

def sort_key_en_index(e: dict):
    name = (e.get("name_en") or e.get("name") or e.get("_id") or "").strip()
    return (name.lower(), str(e.get("_id","")))

# ===== HTML templates =====
def ko_index_header(total: int) -> str:
    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>GPT Catalog (KO)</title>
  <style>
    :root {{
      --bg:#f6f7fb; --card:#fff; --text:#111827; --muted:#6b7280; --border:#e5e7eb;
      --primary:#111827; --primaryText:#fff; --chip:#f3f4f6;
    }}
    *{{box-sizing:border-box}}
    body{{margin:0;font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;background:var(--bg);color:var(--text)}}
    .wrap{{max-width:980px;margin:0 auto;padding:24px 16px 56px}}
    .hero{{background:linear-gradient(180deg,#fff,#ffffff00);border:1px solid var(--border);border-radius:18px;padding:18px 18px 14px;box-shadow:0 6px 22px rgba(17,24,39,0.06)}}
    .chips{{display:flex;gap:8px;flex-wrap:wrap;align-items:center}}
    .chip{{display:inline-flex;align-items:center;gap:6px;padding:6px 10px;border-radius:999px;background:var(--chip);color:var(--text);font-size:12px;text-decoration:none;border:1px solid var(--border)}}
    h1{{margin:0;font-size:28px;letter-spacing:-0.02em}}
    .sub{{margin:6px 0 0;color:var(--muted);font-size:14px;line-height:1.5}}
    .search{{margin-top:12px;display:flex;gap:8px;flex-wrap:wrap}}
    .search input{{flex:1 1 240px;padding:10px 12px;border-radius:12px;border:1px solid var(--border);outline:none;font-size:13px;background:#fff}}
    .hint{{margin-top:8px;color:var(--muted);font-size:12px}}
    .grid{{margin-top:16px;display:grid;grid-template-columns:1fr;gap:12px}}
    .item{{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:14px 14px 12px;box-shadow:0 6px 20px rgba(17,24,39,0.04)}}
    .title{{display:flex;gap:10px;align-items:baseline;justify-content:space-between;flex-wrap:wrap}}
    .name{{font-size:16px;font-weight:750;letter-spacing:-0.01em;margin:0}}
    .meta{{margin-top:6px;color:var(--muted);font-size:13px;line-height:1.5}}
    .btnbar{{margin-top:10px;display:flex;gap:8px;flex-wrap:wrap;align-items:center}}
    .btn{{display:inline-flex;align-items:center;justify-content:center;gap:8px;padding:8px 12px;border-radius:12px;border:1px solid var(--border);text-decoration:none;color:var(--text);font-size:12px;background:#fff}}
    .btn.primary{{background:var(--primary);color:var(--primaryText);border-color:var(--primary)}}
    .badge{{display:inline-flex;align-items:center;padding:6px 10px;border-radius:999px;font-size:12px;border:1px solid var(--border);background:#fff3cd;color:#7a5a00}}
    .titlelink{{color:inherit;text-decoration:none}}
    .titlelink:hover{{text-decoration:underline}}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="hero">
      <div class="chips">
        <span class="chip"><strong>Total</strong> {total}</span>
        <a class="chip" href="en/index.html">English</a>
        <a class="chip" href="{esc(TISTORY_BACK_DEFAULT)}" target="_blank" rel="noopener noreferrer">티스토리 목록</a>
      </div>
      <h1>GPT Catalog (KO)</h1>
      <p class="sub">카탈로그 목록입니다. 항목을 클릭하면 상세 페이지로 이동합니다.</p>
      <div class="search">
        <input id="q" type="search" placeholder="이름/태그/설명 검색…" oninput="filterList()" />
      </div>
      <div class="hint">Tip: 검색어를 입력하면 이름/태그/설명에서 함께 검색됩니다.</div>
    </div>

    <div class="grid" id="list">
"""

def en_index_header(total: int) -> str:
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>GPT Catalog (EN)</title>
  <style>
    :root {{
      --bg:#f6f7fb; --card:#fff; --text:#111827; --muted:#6b7280; --border:#e5e7eb;
      --primary:#111827; --primaryText:#fff; --chip:#f3f4f6;
    }}
    *{{box-sizing:border-box}}
    body{{margin:0;font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;background:var(--bg);color:var(--text)}}
    .wrap{{max-width:980px;margin:0 auto;padding:24px 16px 56px}}
    .hero{{background:linear-gradient(180deg,#fff,#ffffff00);border:1px solid var(--border);border-radius:18px;padding:18px 18px 14px;box-shadow:0 6px 22px rgba(17,24,39,0.06)}}
    .chips{{display:flex;gap:8px;flex-wrap:wrap;align-items:center}}
    .chip{{display:inline-flex;align-items:center;gap:6px;padding:6px 10px;border-radius:999px;background:var(--chip);color:var(--text);font-size:12px;text-decoration:none;border:1px solid var(--border)}}
    h1{{margin:0;font-size:28px;letter-spacing:-0.02em}}
    .sub{{margin:6px 0 0;color:var(--muted);font-size:14px;line-height:1.5}}
    .search{{margin-top:12px;display:flex;gap:8px;flex-wrap:wrap}}
    .search input{{flex:1 1 240px;padding:10px 12px;border-radius:12px;border:1px solid var(--border);outline:none;font-size:13px;background:#fff}}
    .hint{{margin-top:8px;color:var(--muted);font-size:12px}}
    .grid{{margin-top:16px;display:grid;grid-template-columns:1fr;gap:12px}}
    .item{{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:14px 14px 12px;box-shadow:0 6px 20px rgba(17,24,39,0.04)}}
    .title{{display:flex;gap:10px;align-items:baseline;justify-content:space-between;flex-wrap:wrap}}
    .name{{font-size:16px;font-weight:750;letter-spacing:-0.01em;margin:0}}
    .meta{{margin-top:6px;color:var(--muted);font-size:13px;line-height:1.5}}
    .btnbar{{margin-top:10px;display:flex;gap:8px;flex-wrap:wrap;align-items:center}}
    .btn{{display:inline-flex;align-items:center;justify-content:center;gap:8px;padding:8px 12px;border-radius:12px;border:1px solid var(--border);text-decoration:none;color:var(--text);font-size:12px;background:#fff}}
    .btn.primary{{background:var(--primary);color:var(--primaryText);border-color:var(--primary)}}
    .badge{{display:inline-flex;align-items:center;padding:6px 10px;border-radius:999px;font-size:12px;border:1px solid var(--border);background:#fff3cd;color:#7a5a00}}
    .titlelink{{color:inherit;text-decoration:none}}
    .titlelink:hover{{text-decoration:underline}}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="hero">
      <div class="chips">
        <span class="chip"><strong>Total</strong> {total}</span>
        <a class="chip" href="../index.html">Korean</a>
        <a class="chip" href="{esc(TISTORY_BACK_DEFAULT)}" target="_blank" rel="noopener noreferrer">Tistory</a>
      </div>
      <h1>GPT Catalog (EN)</h1>
      <p class="sub">Browse the catalog. Click an item to open its detail page.</p>
      <div class="search">
        <input id="q" type="search" placeholder="Search name/tags/description…" oninput="filterList()" />
      </div>
      <div class="hint">Tip: Search matches name, tags, and description.</div>
    </div>

    <div class="grid" id="list">
"""

def index_footer() -> str:
    return """
    </div>

    <script>
      function norm(s){ return (s||"").toString().toLowerCase(); }
      function filterList(){
        var q = norm(document.getElementById("q").value);
        var items = document.querySelectorAll(".item");
        items.forEach(function(it){
          var hay = norm(it.getAttribute("data-name")) + " " + norm(it.getAttribute("data-tags")) + " " + norm(it.getAttribute("data-one"));
          it.style.display = (q && hay.indexOf(q) === -1) ? "none" : "";
        });
      }
    </script>
  </div>
</body>
</html>
"""

def make_detail_link(e: dict, lang: str) -> str:
    if lang == "ko":
        return f"details/{e['_id']}_ko.html"
    return f"../details/{e['_id']}_en.html"

def make_back_link_from_detail(lang: str) -> str:
    return "../index.html" if lang == "ko" else "../en/index.html"

def detail_page(e: dict, lang: str) -> str:
    is_ko = (lang == "ko")
    title_text = (e.get("name_ko") or "").strip() if is_ko else (e.get("name_en") or e.get("name") or "").strip()
    if not title_text:
        title_text = (e.get("name_en") or e.get("name") or e.get("_id") or "").strip()
    subtitle = (e.get("one_line") or "").strip()

    back_href = make_back_link_from_detail(lang)
    back_label = "← 목록" if is_ko else "← Back"
    primary_label = "GPT 바로가기" if is_ko else "Open GPT"
    tistory_label = "티스토리" if is_ko else "Tistory"

    show_url = (e.get("viz2", DEFAULT_VIZ2) != "hide_url")
    restricted = (e.get("viz1", DEFAULT_VIZ1) == "restricted")
    gpt_url = (e.get("url") or "").strip()

    funcs = e.get("functions") or []
    if isinstance(funcs, str):
        funcs = [funcs]
    func_items = "".join(f"<li class='fitem'>{esc(x)}</li>" for x in funcs if x)

    tech_rows = [
        ("_id", e.get("_id","")),
        ("gpt_id", e.get("gpt_id","")),
        ("version", e.get("version","")),
        ("last_updated", e.get("last_updated","")),
        ("viz1", e.get("viz1", DEFAULT_VIZ1)),
        ("viz2", e.get("viz2", DEFAULT_VIZ2)),
        ("url", gpt_url),
    ]
    tech_html = "".join(
        f"<div class='row'><div class='k'>{esc(k)}</div><div class='v'>{esc(v)}</div></div>"
        for k, v in tech_rows
        if v not in (None, "")
    )

    notice = ""
    if restricted:
        notice = "<div class='notice'>" + ("제한공개 항목입니다." if is_ko else "This item is restricted.") + "</div>"

    if gpt_url and show_url and not restricted:
        primary_btn = f"<a class='btn primary' href='{esc(gpt_url)}' target='_blank' rel='noopener noreferrer'>{esc(primary_label)}</a>"
    elif restricted:
        primary_btn = f"<span class='btn disabled'>{'제한공개' if is_ko else 'Restricted'}</span>"
    elif gpt_url and not show_url:
        primary_btn = f"<span class='btn disabled'>{'URL 숨김' if is_ko else 'URL hidden'}</span>"
    else:
        primary_btn = ""

    return f"""<!doctype html>
<html lang="{ 'ko' if is_ko else 'en' }">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(title_text)}</title>
  <style>
    :root {{
      --bg:#f6f7fb; --card:#fff; --text:#111827; --muted:#6b7280; --border:#e5e7eb;
      --primary:#111827; --primaryText:#fff;
    }}
    *{{box-sizing:border-box}}
    body{{margin:0;font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;background:var(--bg);color:var(--text)}}
    .wrap{{max-width:980px;margin:0 auto;padding:24px 16px 56px}}
    .card{{background:var(--card);border:1px solid var(--border);border-radius:18px;padding:18px;box-shadow:0 6px 22px rgba(17,24,39,0.06)}}
    .topbar{{display:flex;gap:10px;flex-wrap:wrap;align-items:center;justify-content:space-between;margin-bottom:12px}}
    .leftbar{{display:flex;gap:8px;flex-wrap:wrap;align-items:center}}
    .btn{{display:inline-flex;align-items:center;justify-content:center;gap:8px;padding:10px 14px;border-radius:14px;border:1px solid var(--border);text-decoration:none;color:var(--text);font-size:13px;background:#fff}}
    .btn.primary{{background:var(--primary);color:var(--primaryText);border-color:var(--primary);font-weight:700}}
    .btn.disabled{{background:#f3f4f6;color:#6b7280;border-color:var(--border);cursor:default}}
    h1{{margin:0;font-size:30px;letter-spacing:-0.02em}}
    .sub{{margin:10px 0 0;color:var(--muted);line-height:1.6;font-size:14px}}
    .notice{{margin-top:12px;padding:10px 12px;border-radius:14px;border:1px solid var(--border);background:#fff3cd;color:#7a5a00;font-size:13px}}
    .section{{margin-top:14px}}
    .h2{{margin:0 0 8px;font-size:16px;letter-spacing:-0.01em}}
    ul.funcs{{list-style:none;padding:0;margin:0;display:grid;grid-template-columns:1fr;gap:10px}}
    .fitem{{background:#fff;border:1px solid var(--border);border-radius:16px;padding:12px 12px;line-height:1.55;box-shadow:0 6px 20px rgba(17,24,39,0.04);font-size:14px}}
    details{{margin-top:14px;border:1px solid var(--border);border-radius:16px;background:#fff;padding:10px 12px}}
    summary{{cursor:pointer;font-weight:700;color:var(--text);list-style:none;outline:none}}
    summary::-webkit-details-marker{{display:none}}
    .kv{{margin-top:10px;display:grid;gap:8px}}
    .row{{display:grid;grid-template-columns:140px 1fr;gap:10px;padding:8px 10px;border-radius:12px;background:#f9fafb;border:1px solid var(--border);overflow-wrap:anywhere}}
    .k{{color:var(--muted);font-size:12px}}
    .v{{color:var(--text);font-size:13px}}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div class="topbar">
        <div class="leftbar">
          <a class="btn" href="{esc(back_href)}">{esc(back_label)}</a>
          {primary_btn}
          <a class="btn" href="{esc(TISTORY_BACK_DEFAULT)}" target="_blank" rel="noopener noreferrer">{esc(tistory_label)}</a>
        </div>
      </div>

      <h1>{esc(title_text)}</h1>
      {f"<p class='sub'>{esc(subtitle)}</p>" if subtitle else ""}
      {notice}

      <div class="section">
        <div class="h2">{'핵심 기능' if is_ko else 'Key functions'}</div>
        {("<ul class='funcs'>"+func_items+"</ul>") if func_items else ("<div class='sub'>"+("등록된 기능이 없습니다." if is_ko else "No functions listed.")+"</div>")}
      </div>

      <details>
        <summary>{'기술 정보' if is_ko else 'Technical details'}</summary>
        <div class="kv">{tech_html}</div>
      </details>
    </div>
  </div>
</body>
</html>
"""

def build(clean: bool = True):
    if clean:
        safe_rmtree(DETAILS_DIR)
        safe_rmtree(EN_DIR)
        root_index = os.path.join(OUT_DIR, "index.html")
        if os.path.isfile(root_index):
            try:
                os.remove(root_index)
            except Exception:
                pass

    os.makedirs(DETAILS_DIR, exist_ok=True)
    os.makedirs(EN_DIR, exist_ok=True)

    index_en, index_ko = load_excel_index()

    all_entries = []
    unknown_files = []

    for fn in list_catalog_files():
        path = os.path.join(CATALOG_DIR, fn)
        raw = read_yaml(path)
        if raw is None or not isinstance(raw, dict):
            continue

        e = raw.get("catalog_entry") or raw.get("gpt") or raw

        raw_id = e.get("_id") or e.get("id") or e.get("gpt_id") or os.path.splitext(fn)[0]
        e["_id"] = sanitize_id(str(raw_id))

        e["_lang"] = guess_lang(e, fn)
        if e["_lang"] == "unknown":
            unknown_files.append((fn, e.get("_id","")))

        apply_excel_override(e, index_en, index_ko)
        all_entries.append(e)

    ko_entries = [e for e in all_entries if e["_lang"] == "ko"]
    en_entries = [e for e in all_entries if e["_lang"] == "en"]
    unknown_entries = [e for e in all_entries if e["_lang"] == "unknown"]

    ko_entries.sort(key=sort_key_ko_index)
    en_entries.sort(key=sort_key_en_index)

    print(f"Loaded total: {len(all_entries)} | KO: {len(ko_entries)} | EN: {len(en_entries)}")
    if unknown_entries:
        print(f"[WARN] Unknown-language catalogs skipped: {len(unknown_entries)}")
        for fn, _id in unknown_files[:50]:
            print(f"       - {fn} (id={_id})")

    # KO index
    root_index = os.path.join(OUT_DIR, "index.html")
    with open(root_index, "w", encoding="utf-8") as f:
        f.write(ko_index_header(len(ko_entries)))
        for i, e in enumerate(ko_entries, start=1):
            name = display_name_ko(e)
            one = esc(e.get("one_line", ""))
            tags_raw = e.get("tags", [])
            tags = esc(" ".join(map(str, tags_raw))) if isinstance(tags_raw, list) else esc(str(tags_raw))
            gpt_url = esc(e.get("url", ""))
            detail_href = make_detail_link(e, "ko")

            restricted = (e.get("viz1", DEFAULT_VIZ1) == "restricted")
            show_url = (e.get("viz2", DEFAULT_VIZ2) != "hide_url")

            if restricted:
                title_href = "#"
                detail_btn = "<span class='badge'>제한공개</span>"
            else:
                title_href = detail_href
                detail_btn = f"<a class='btn' href='{detail_href}'>상세보기</a>"

            url_btn = ""
            if gpt_url and show_url and not restricted:
                url_btn = f"<a class='btn primary' href='{gpt_url}' target='_blank' rel='noopener noreferrer'>GPT 바로가기</a>"

            f.write(
                f"<div class='item' data-name='{esc(_primary_display_ko(e))}' data-tags='{tags}' data-one='{one}'>"
                f"<div class='title'>"
                f"<a class='titlelink' href='{title_href}'><div class='name'>{i}. {name}</div></a>"
                f"</div>"
                f"<div class='meta'>{one}</div>"
                f"<div class='btnbar'>{url_btn}{detail_btn}</div>"
                f"</div>"
            )
        f.write(index_footer())

    # EN index
    en_index = os.path.join(EN_DIR, "index.html")
    with open(en_index, "w", encoding="utf-8") as f:
        f.write(en_index_header(len(en_entries)))
        for i, e in enumerate(en_entries, start=1):
            name = display_name_en(e)
            one = esc(e.get("one_line", ""))
            tags_raw = e.get("tags", [])
            tags = esc(" ".join(map(str, tags_raw))) if isinstance(tags_raw, list) else esc(str(tags_raw))
            gpt_url = esc(e.get("url", ""))
            detail_href = make_detail_link(e, "en")

            restricted = (e.get("viz1", DEFAULT_VIZ1) == "restricted")
            show_url = (e.get("viz2", DEFAULT_VIZ2) != "hide_url")

            if restricted:
                title_href = "#"
                detail_btn = "<span class='badge'>Restricted</span>"
            else:
                title_href = detail_href
                detail_btn = f"<a class='btn' href='{detail_href}'>View details</a>"

            url_btn = ""
            if gpt_url and show_url and not restricted:
                url_btn = f"<a class='btn primary' href='{gpt_url}' target='_blank' rel='noopener noreferrer'>Open GPT</a>"

            f.write(
                f"<div class='item' data-name='{esc(e.get('name_en') or e.get('name') or e.get('_id') or '')}' "
                f"data-tags='{tags}' data-one='{one}'>"
                f"<div class='title'>"
                f"<a class='titlelink' href='{title_href}'><div class='name'>{i}. {name}</div></a>"
                f"</div>"
                f"<div class='meta'>{one}</div>"
                f"<div class='btnbar'>{url_btn}{detail_btn}</div>"
                f"</div>"
            )
        f.write(index_footer())

    # Details pages
    for e in ko_entries:
        out = os.path.join(DETAILS_DIR, f"{e['_id']}_ko.html")
        with open(out, "w", encoding="utf-8") as fp:
            fp.write(detail_page(e, "ko"))
    for e in en_entries:
        out = os.path.join(DETAILS_DIR, f"{e['_id']}_en.html")
        with open(out, "w", encoding="utf-8") as fp:
            fp.write(detail_page(e, "en"))

    print(f"[OK] GitHub docs written: {OUT_DIR}")
    print(f"[OK] Details: {DETAILS_DIR}")

if __name__ == "__main__":
    build(clean=True)